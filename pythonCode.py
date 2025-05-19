import serial
import time
from openpyxl import Workbook, load_workbook
import os

# إنشاء أو تحميل ملف Excel
file_name = "sensor_readings.xlsx"
if os.path.exists(file_name):
    workbook = load_workbook(file_name)
    sheet = workbook.active
else:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sensor Readings"
    sheet.append(["Timestamp", "Reading"])  # العناوين

def check_for_warnings(file_name):
    if os.path.exists(file_name):
        workbook = load_workbook(file_name)
        sheet = workbook.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[1]:
                values = row[1].split("|")
                for value in values:
                    if ":" in value:
                        _, val = value.split(":")
                        if val.strip().isdigit() and int(val.strip()) > 0:
                            print("\n⚠️ Warning: Detected values above 0 in Excel sheet.")
                            return
        print("\n✅ No warnings. All readings are safe.")

try:
    ser = serial.Serial('COM2', 9600)  # COM2 لأن COM1 مربوط مع Proteus
    time.sleep(2)  # انتظار لبدء الاتصال

    print("Listening for data from Proteus...")

    while True:
        if ser.in_waiting > 0:  # التحقق من وجود بيانات في الـSerial
            data = ser.readline().decode(errors='ignore').strip()  # قراءة البيانات وفك تشفيرها
            if data:
                timestamp = time.strftime("%Y-%m-%d %H:%M:%S")
                print(f"Received from Proteus: {data}")

                # حفظ القراءة في Excel
                sheet.append([timestamp, data])
                workbook.save(file_name)

except serial.SerialException as e:
    print(f"[!] Serial error: {e}")
except KeyboardInterrupt:
    print("\n[!] Program interrupted by user.")
    print("\n[!] Checking for warnings in Excel before exit...")
    check_for_warnings(file_name)
finally:
    ser.close()
    workbook.save(file_name)
    print(f"\n[!] Data saved to {file_name}")
